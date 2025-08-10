import json
import os
from typing import List, Dict, Any, Optional
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from google import genai
from google.genai import types
import base64
from io import BytesIO
from flask import Flask, request, jsonify, render_template_string
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

class ExcelProcessor:
    """Efficient Excel file processor optimized for LLM token usage"""
    
    def __init__(self):
        self.max_rows_per_sheet = 100  # Limit rows to control token usage
        self.max_chars_per_cell = 500  # Limit cell content length
        
    def read_excel_file(self, file_path: str, file_name: str) -> Dict[str, Any]:
        """Read Excel file and extract structured data"""
        try:
            # Load workbook
            workbook = load_workbook(file_path, data_only=True)
            
            file_data = {
                'file_name': file_name,
                'sheets': {},
                'summary': {
                    'total_sheets': len(workbook.sheetnames),
                    'sheet_names': workbook.sheetnames
                }
            }
            
            for sheet_name in workbook.sheetnames:
                sheet_data = self._process_sheet(workbook[sheet_name], sheet_name)
                file_data['sheets'][sheet_name] = sheet_data
                
            return file_data
            
        except Exception as e:
            logger.error(f"Error reading Excel file {file_name}: {str(e)}")
            return {'error': f"Failed to read {file_name}: {str(e)}"}
    
    def _process_sheet(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Process individual sheet with optimization for LLM"""
        try:
            # Get sheet dimensions
            max_row = min(sheet.max_row, self.max_rows_per_sheet + 1)  # +1 for header
            max_col = sheet.max_column
            
            # Extract headers (first row)
            headers = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")
            
            # Extract data rows
            data_rows = []
            for row in range(2, max_row + 1):
                row_data = {}
                has_data = False
                
                for col, header in enumerate(headers, 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    
                    # Clean and limit cell content
                    if cell_value is not None:
                        cell_str = str(cell_value)
                        if len(cell_str) > self.max_chars_per_cell:
                            cell_str = cell_str[:self.max_chars_per_cell] + "..."
                        row_data[header] = cell_str
                        has_data = True
                    else:
                        row_data[header] = ""
                
                if has_data:
                    data_rows.append(row_data)
            
            # Create summary statistics
            summary = {
                'total_rows': len(data_rows),
                'total_columns': len(headers),
                'headers': headers,
                'data_types': self._analyze_data_types(data_rows, headers),
                'sample_data': data_rows[:3] if data_rows else []  # First 3 rows as sample
            }
            
            return {
                'sheet_name': sheet_name,
                'summary': summary,
                'data': data_rows
            }
            
        except Exception as e:
            logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
            return {'error': f"Failed to process sheet {sheet_name}: {str(e)}"}
    
    def _analyze_data_types(self, data_rows: List[Dict], headers: List[str]) -> Dict[str, str]:
        """Analyze data types for each column"""
        data_types = {}
        
        for header in headers:
            sample_values = [row.get(header, "") for row in data_rows[:10] if row.get(header, "")]
            
            if not sample_values:
                data_types[header] = "empty"
                continue
                
            # Simple type detection
            numeric_count = sum(1 for val in sample_values if str(val).replace('.', '').replace('-', '').isdigit())
            
            if numeric_count > len(sample_values) * 0.7:
                data_types[header] = "numeric"
            else:
                data_types[header] = "text"
                
        return data_types
    
    def create_llm_optimized_summary(self, files_data: List[Dict[str, Any]]) -> str:
        """Create a token-efficient summary for LLM"""
        summary_parts = []
        
        summary_parts.append("=== EXCEL FILES ANALYSIS SUMMARY ===")
        summary_parts.append(f"Total files processed: {len(files_data)}")
        
        for file_data in files_data:
            if 'error' in file_data:
                summary_parts.append(f"\n❌ {file_data.get('file_name', 'Unknown')}: {file_data['error']}")
                continue
                
            file_name = file_data['file_name']
            summary_parts.append(f"\n📁 FILE: {file_name}")
            summary_parts.append(f"   Sheets: {file_data['summary']['total_sheets']} ({', '.join(file_data['summary']['sheet_names'])})")
            
            for sheet_name, sheet_data in file_data['sheets'].items():
                if 'error' in sheet_data:
                    summary_parts.append(f"   ❌ Sheet '{sheet_name}': {sheet_data['error']}")
                    continue
                    
                summary = sheet_data['summary']
                summary_parts.append(f"\n   📊 SHEET: {sheet_name}")
                summary_parts.append(f"      Dimensions: {summary['total_rows']} rows × {summary['total_columns']} columns")
                summary_parts.append(f"      Columns: {', '.join(summary['headers'][:10])}{'...' if len(summary['headers']) > 10 else ''}")
                
                # Add sample data
                if summary['sample_data']:
                    summary_parts.append("      Sample data:")
                    for i, row in enumerate(summary['sample_data'][:2], 1):
                        row_preview = {k: str(v)[:50] + ("..." if len(str(v)) > 50 else "") for k, v in row.items()}
                        summary_parts.append(f"        Row {i}: {json.dumps(row_preview, ensure_ascii=False)}")
        
        return "\n".join(summary_parts)

class GeminiLLM:
    """Gemini LLM integration"""
    
    def __init__(self, api_key: str):
        self.client = genai.Client(api_key=api_key)
        self.model = "gemini-2.0-flash"
    
    def analyze_excel_data(self, excel_summary: str, user_query: str = "") -> str:
        """Send Excel data to Gemini for analysis"""
        try:
            prompt = f"""
You are an expert data analyst. I have processed multiple Excel files and need your analysis.

EXCEL DATA SUMMARY:
{excel_summary}

USER QUERY: {user_query if user_query else "Please provide a comprehensive analysis of this Excel data, including key insights, patterns, and recommendations."}

Please provide:
1. Key insights from the data
2. Data quality observations
3. Patterns or trends you notice
4. Recommendations for further analysis
5. Any potential issues or anomalies

Be specific and actionable in your response.
"""

            contents = [
                types.Content(
                    role="user",
                    parts=[types.Part.from_text(text=prompt)],
                ),
            ]
            
            tools = [types.Tool(googleSearch=types.GoogleSearch())]
            
            generate_content_config = types.GenerateContentConfig(tools=tools)
            
            response_text = ""
            for chunk in self.client.models.generate_content_stream(
                model=self.model,
                contents=contents,
                config=generate_content_config,
            ):
                response_text += chunk.text
                
            return response_text
            
        except Exception as e:
            logger.error(f"Error calling Gemini API: {str(e)}")
            return f"Error analyzing data with Gemini: {str(e)}"

# Global instances
excel_processor = ExcelProcessor()
gemini_llm = None

# HTML Template
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel to LLM Processor</title>
    <style>
        body { font-family: Arial, sans-serif; max-width: 1200px; margin: 0 auto; padding: 20px; }
        .container { background: #f5f5f5; padding: 20px; border-radius: 8px; margin-bottom: 20px; }
        .upload-area { border: 2px dashed #ccc; padding: 40px; text-align: center; border-radius: 8px; }
        .upload-area.dragover { border-color: #007bff; background: #e3f2fd; }
        button { background: #007bff; color: white; padding: 12px 24px; border: none; border-radius: 4px; cursor: pointer; margin: 5px; }
        button:hover { background: #0056b3; }
        button:disabled { background: #6c757d; cursor: not-allowed; }
        .file-list { margin: 20px 0; }
        .file-item { background: white; padding: 10px; margin: 5px 0; border-radius: 4px; border-left: 4px solid #007bff; }
        .result { background: white; padding: 20px; border-radius: 8px; margin-top: 20px; }
        .loading { text-align: center; padding: 20px; }
        .error { color: #dc3545; }
        .success { color: #28a745; }
        textarea { width: 100%; min-height: 100px; padding: 10px; border-radius: 4px; border: 1px solid #ddd; }
        pre { background: #f8f9fa; padding: 15px; border-radius: 4px; overflow-x: auto; white-space: pre-wrap; }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 Excel to LLM Processor</h1>
        <p>Upload multiple Excel files to analyze with AI. Optimized for efficient token usage.</p>
    </div>

    <div class="container">
        <h2>🔑 Gemini API Key</h2>
        <input type="text" id="apiKey" placeholder="Enter your Gemini API key" style="width: 100%; padding: 10px; margin-bottom: 10px;">
        <button onclick="setApiKey()">Set API Key</button>
        <div id="apiKeyStatus"></div>
    </div>

    <div class="container">
        <h2>📁 Upload Excel Files</h2>
        <div class="upload-area" id="uploadArea">
            <p>Drag & drop Excel files here or click to select</p>
            <input type="file" id="fileInput" multiple accept=".xlsx,.xls" style="display: none;">
            <button onclick="document.getElementById('fileInput').click()">Select Files</button>
        </div>
        <div class="file-list" id="fileList"></div>
    </div>

    <div class="container">
        <h2>❓ Your Question (Optional)</h2>
        <textarea id="userQuery" placeholder="Ask a specific question about your Excel data, or leave blank for general analysis..."></textarea>
    </div>

    <div class="container">
        <button id="processBtn" onclick="processFiles()" disabled>🚀 Process & Analyze</button>
        <button onclick="clearAll()">🗑️ Clear All</button>
    </div>

    <div id="results"></div>

    <script>
        let selectedFiles = [];
        let apiKey = '';

        // API Key management
        function setApiKey() {
            const key = document.getElementById('apiKey').value.trim();
            if (key) {
                apiKey = key;
                document.getElementById('apiKeyStatus').innerHTML = '<span class="success">✅ API Key set successfully</span>';
                updateProcessButton();
            } else {
                document.getElementById('apiKeyStatus').innerHTML = '<span class="error">❌ Please enter a valid API key</span>';
            }
        }

        // File upload handling
        const uploadArea = document.getElementById('uploadArea');
        const fileInput = document.getElementById('fileInput');

        uploadArea.addEventListener('click', () => fileInput.click());
        uploadArea.addEventListener('dragover', (e) => {
            e.preventDefault();
            uploadArea.classList.add('dragover');
        });
        uploadArea.addEventListener('dragleave', () => {
            uploadArea.classList.remove('dragover');
        });
        uploadArea.addEventListener('drop', (e) => {
            e.preventDefault();
            uploadArea.classList.remove('dragover');
            handleFiles(e.dataTransfer.files);
        });

        fileInput.addEventListener('change', (e) => {
            handleFiles(e.target.files);
        });

        function handleFiles(files) {
            for (let file of files) {
                if (file.name.endsWith('.xlsx') || file.name.endsWith('.xls')) {
                    selectedFiles.push(file);
                }
            }
            displayFileList();
            updateProcessButton();
        }

        function displayFileList() {
            const fileList = document.getElementById('fileList');
            if (selectedFiles.length === 0) {
                fileList.innerHTML = '';
                return;
            }

            fileList.innerHTML = '<h3>Selected Files:</h3>' + 
                selectedFiles.map((file, index) => 
                    `<div class="file-item">
                        📄 ${file.name} (${(file.size / 1024 / 1024).toFixed(2)} MB)
                        <button onclick="removeFile(${index})" style="float: right; background: #dc3545;">Remove</button>
                    </div>`
                ).join('');
        }

        function removeFile(index) {
            selectedFiles.splice(index, 1);
            displayFileList();
            updateProcessButton();
        }

        function updateProcessButton() {
            const processBtn = document.getElementById('processBtn');
            processBtn.disabled = !(selectedFiles.length > 0 && apiKey);
        }

        function clearAll() {
            selectedFiles = [];
            document.getElementById('fileList').innerHTML = '';
            document.getElementById('results').innerHTML = '';
            document.getElementById('userQuery').value = '';
            updateProcessButton();
        }

        async function processFiles() {
            if (selectedFiles.length === 0 || !apiKey) return;

            const results = document.getElementById('results');
            results.innerHTML = '<div class="loading">🔄 Processing files and analyzing with AI...</div>';

            const formData = new FormData();
            formData.append('api_key', apiKey);
            formData.append('user_query', document.getElementById('userQuery').value);
            
            selectedFiles.forEach(file => {
                formData.append('files', file);
            });

            try {
                const response = await fetch('/api/process', {
                    method: 'POST',
                    body: formData
                });

                const data = await response.json();

                if (data.success) {
                    results.innerHTML = `
                        <div class="result">
                            <h2>✅ Processing Complete</h2>
                            <h3>📊 Data Summary:</h3>
                            <pre>${data.excel_summary}</pre>
                            <h3>🤖 AI Analysis:</h3>
                            <pre>${data.llm_analysis}</pre>
                        </div>
                    `;
                } else {
                    results.innerHTML = `
                        <div class="result">
                            <h2 class="error">❌ Processing Failed</h2>
                            <p>${data.error}</p>
                        </div>
                    `;
                }
            } catch (error) {
                results.innerHTML = `
                    <div class="result">
                        <h2 class="error">❌ Network Error</h2>
                        <p>Failed to process files: ${error.message}</p>
                    </div>
                `;
            }
        }
    </script>
</body>
</html>
"""

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/api/process', methods=['POST'])
def process_excel_files():
    global gemini_llm
    
    try:
        # Get API key and user query
        api_key = request.form.get('api_key')
        user_query = request.form.get('user_query', '')
        
        if not api_key:
            return jsonify({'success': False, 'error': 'API key is required'})
        
        # Initialize Gemini LLM
        gemini_llm = GeminiLLM(api_key)
        
        # Get uploaded files
        files = request.files.getlist('files')
        if not files:
            return jsonify({'success': False, 'error': 'No files uploaded'})
        
        # Process each Excel file
        files_data = []
        temp_files = []
        
        for file in files:
            if file and file.filename.endswith(('.xlsx', '.xls')):
                # Save temporary file
                temp_path = f"/tmp/{file.filename}"
                file.save(temp_path)
                temp_files.append(temp_path)
                
                # Process Excel file
                file_data = excel_processor.read_excel_file(temp_path, file.filename)
                files_data.append(file_data)
        
        # Create LLM-optimized summary
        excel_summary = excel_processor.create_llm_optimized_summary(files_data)
        
        # Get AI analysis
        llm_analysis = gemini_llm.analyze_excel_data(excel_summary, user_query)
        
        # Clean up temporary files
        for temp_file in temp_files:
            try:
                os.remove(temp_file)
            except:
                pass
        
        return jsonify({
            'success': True,
            'excel_summary': excel_summary,
            'llm_analysis': llm_analysis,
            'files_processed': len(files_data)
        })
        
    except Exception as e:
        logger.error(f"Error processing files: {str(e)}")
        return jsonify({'success': False, 'error': str(e)})

# Vercel serverless function handler
def handler(request):
    return app(request.environ, lambda s, h: None)

if __name__ == '__main__':
    app.run(debug=True)
